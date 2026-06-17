"""
Extrator de escopo de planilhas .xlsx para o agente MUD Engenharia.

Lê todas as abas, preserva cabeçalhos e linhas em formato legível.
Ignora linhas totalmente vazias e células com fórmula (data_only=True).
"""

import io
from typing import Any

import openpyxl


def _celula_str(valor: Any) -> str:
    if valor is None:
        return ""
    return str(valor).strip()


def extrair_planilha(conteudo: bytes, nome_arquivo: str = "planilha.xlsx") -> tuple[str, list[str]]:
    """
    Extrai o texto de todas as abas de uma planilha .xlsx.
    Retorna (texto_escopo, avisos).
    """
    avisos: list[str] = []
    partes: list[str] = []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True, read_only=True)
    except Exception as e:
        msg = f"Erro ao abrir planilha '{nome_arquivo}': {e}"
        avisos.append(msg)
        return "", avisos

    for nome_aba in wb.sheetnames:
        try:
            ws = wb[nome_aba]
            linhas_aba: list[str] = []

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                celulas = [_celula_str(c) for c in row]

                # Ignora linha totalmente vazia
                if not any(celulas):
                    continue

                linha_str = " | ".join(celulas)

                # Primeira linha não vazia vira cabeçalho
                if not linhas_aba:
                    linhas_aba.append(f"CABEÇALHO: {linha_str}")
                else:
                    linhas_aba.append(linha_str)

            if linhas_aba:
                partes.append(f"[Aba: {nome_aba}]")
                partes.extend(linhas_aba)
            else:
                avisos.append(f"Aba '{nome_aba}' está vazia e foi ignorada.")

        except Exception as e:
            avisos.append(f"Erro ao ler aba '{nome_aba}' de '{nome_arquivo}': {e}")

    wb.close()

    if not partes:
        avisos.append(f"Planilha '{nome_arquivo}' não contém dados legíveis.")
        return "", avisos

    return "\n".join(partes), avisos
